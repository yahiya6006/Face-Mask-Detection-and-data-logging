''' Step 1: We will have a look on how to create a xl document
    step 2: create a class to form tabels in an xl document
'''

class XLC:

    def create(self, data, name, sheetname):
        """This Method is going to create an XL sheet with first input as a sting
            i,e "Time,Data" etc which will be coloums and the secand input as the
            Filename and the third input as the name of the sheet"""
        # Importing Font to change the font 
        from openpyxl.styles import Font
        self.data = data
        self.name = name + ".xlsx"
        self.sheetname = sheetname

        # Created a tupple that contains all the names of coloums in an XL sheet
        self.colum = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R'
                      ,'S','T','U','V')

        # Checking if file is already available
        self.cf = XLC.checkfile(self,self.name)
        if self.cf == True:
            print("[INFO] File {} already exists".format(self.name))
            userAns = input("[INFO] Do you want to over write the data \n[INFO] over writing the data will erase all the data stored in {} y/n: ".format(self.filename))
            if userAns == "y" or userAns == "Y":
                wb = xl.Workbook()
                wb.sheetnames
                sheet = wb.active
                sheet.title = self.sheetname

                # Seperating the received data using split
                self.data = self.data.split(",")
                for i in range(0, len(self.data)):

                    # When for loop starts we select 1st coloum i.e A1 simillarly B1 C1....
                    # using sheet[self.colum[i]+'1'] and the data is stored in it
                    sheet[self.colum[i]+'1'] = self.data[i]

                    # Setting the Font to Times New Roman 
                    sheet[self.colum[i]+'1'].font = Font(name='Times New Roman', bold=True, size=14)

                    # Setting the coloum width
                    sheet.column_dimensions[self.colum[i]].width = len(self.data[i])+12
                wb.save(self.name)
                print("{} saved".format(self.filename))
                return

            if userAns == "n" or userAns == "N":
                return
            
        if self.cf == False:
            # Creating the workbook
            wb = xl.Workbook()
            wb.sheetnames
            sheet = wb.active
            sheet.title = self.sheetname

            # Seperating the received data using split
            self.data = self.data.split(",")
            for i in range(0, len(self.data)):

                # When for loop starts we select 1st coloum i.e A1 simillarly B1 C1....
                # using sheet[self.colum[i]+'1'] and the data is stored in it
                sheet[self.colum[i]+'1'] = self.data[i]

                # Setting the Font to Times New Roman 
                sheet[self.colum[i]+'1'].font = Font(name='Times New Roman', bold=True, size=14)

                # Setting the coloum width
                sheet.column_dimensions[self.colum[i]].width = len(self.data[i])+12
            wb.save(self.name)
            print("{} saved".format(self.filename))
            return

    def checkfile(self, filename):
        import os
        self.filename = filename
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))

        for root,direct,files in os.walk(BASE_DIR):
            for file in files:
                if file == self.filename:
                    return True
        return False

    def updateData(self, data, filename):
        self.data = data
        #self.filename = filename + ".xlsx"
        self.index = "index.txt"
        self.colum = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R'
                      ,'S','T','U','V')
        cf = XLC.checkfile(self, self.index )
        if cf == True:
            self.resultFile = open(self.index, 'r')
            self.row = int(self.resultFile.read())
            self.resultFile.close()
        if cf == False:
            print("[WARNING] index.txt file is missing. Neglect this message if this program is running 1st time on your system")
            self.row = 2
        
        # Looding the document
        self.filename = filename + ".xlsx"
        wb = xl.load_workbook(self.filename)
        sheet = wb.sheetnames
        sheet = wb.active

        if sheet['A2'].value == None:
            self.row = 2

        self.data = self.data.split(",")
        for i in range(0, len(self.data)):
            sheet[self.colum[i]+str(self.row)] = self.data[i]
            print(self.colum[i]+str(self.row))

        self.row = self.row + 1
        self.resultFile = open('index.txt', 'w')
        self.resultFile.write(str(self.row))
        self.resultFile.close()
        wb.save(self.filename)
        return

import openpyxl as xl

'''## Step 1__________________________________________
# Creating a blank workbook
wb = xl.Workbook()

# Starts with the 1st sheet 'sheet' use print to find out
wb.sheetnames

# The sheet is made active
sheet = wb.active

# Changeing the title of the sheeet to "Face Mask detector"
sheet.title = "Face Mask detector"

sheet['A1'] = "Time"

# Saveing the document  
wb.save("FMD.xlsx")'''

## Step 2_____________________________

XL = XLC()
XL.create("Date,Time,Person Identified,Status", "FMD1", "Face Mask detector")
for i in range(0,2):
    XL.updateData("18/05/2020,11:45,yahiya,No Mask Detected","FMD1")
